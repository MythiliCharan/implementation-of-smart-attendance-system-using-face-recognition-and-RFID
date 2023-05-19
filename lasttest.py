import customtkinter
from datetime import date
from datetime import datetime
import serial
import openpyxl
import cv2
import face_recognition
import numpy as np
from datetime import date
import multiprocessing
from collections import Counter
import queue
from PIL import Image
import yagmail
#yag = yagmail.SMTP("mythilicharanr@gmail.com", oauth2_file="/media/mythilicharan/7A04B24604B204E3/client_secret_426456360610-otlapauhnq334snlsdpmjqmu9oq0t4fb.apps.googleusercontent.com.json")

contents = [
    "/home/mythilicharan/Downloads/datas.xlsx"
]


cam_id=0
port='/dev/ttyUSB0'

manager = multiprocessing.Manager()

today_date = date.today().strftime('%d-%m-%Y')
today_name=day=date.today().strftime("%A")
this_month=date.today().strftime("%B")

min_row=6
max_row=None
today_column=None

no=None #presentcount
ab=None #absentcount
totalratio=0
percent=0


names=[]
regno=[]
RFID=[]
path_names=[]

present_names=manager.list()
present_regno=manager.list()
lastpresdata = manager.list()

lastpresdata.append("n")
lastpresdata.append("n")

lastpresdata.append("n")
lastpresdata.append("n")

absent_names=[]
absent_regno=[]


known_face_encodings = [face_recognition.face_encodings(face_recognition.load_image_file("/home/mythilicharan/face_recognition/examples/111.jpg"))[0]]
known_face_names =["MYTHILI CHARAN"]
 

customtkinter.set_appearance_mode("dark")  # Modes: "System" (standard), "Dark", "Light"
customtkinter.set_default_color_theme("blue")  # Themes: "blue" (standard), "green", "dark-blue"

today_date = date.today().strftime('%d-%m-%Y')
today_name=day=date.today().strftime("%A")
this_month=date.today().strftime("%B")
now = datetime.now()

#arduino = serial.Serial(port,115200, timeout=.1)



def excel_update(put_pres):

    daily=openpyxl.load_workbook(outpath+today_date+today_name+"11.xlsx")
    monthly=openpyxl.load_workbook(outpath+this_month+"111.xlsx")
    monthlydata=openpyxl.load_workbook(outpath+this_month+"111.xlsx", data_only=True)

    lastpresdata[2]= now.strftime("%H:%M")
    day=daily.active
    month=monthly.active
    monthdat=monthlydata.active
    for i in range(min_row,max_row):
        cell=day.cell(row=i,column=1)
        print(cell.value)
        if cell.value==put_pres:
            day.cell(row=i,column=3).value=1
            month.cell(row=i,column=today_column).value=1
            lastpresdata[3] = str(monthdat.cell(row=i,column=6).value)
            break
    daily.save(filename=outpath+today_date+today_name+"11.xlsx")
    monthly.save(filename=outpath+this_month+"111.xlsx")


video_capture = cv2.VideoCapture(cam_id)


img_path="/home/mythilicharan/Desktop/project/face data/"
day_excel_path="/home/mythilicharan/Desktop/project/excelsheets/DAILY REPORT.xlsx"
month_excel_path="/home/mythilicharan/Desktop/project/excelsheets/MONTHLY REPORT.xlsx"
outpath="/home/mythilicharan/Music/"


def visionmode(present_names,present_regno):

    print(names)

    

    print("VISION MODE ACTIVATED")
    print("STARTED")
    video_capture = cv2.VideoCapture(cam_id)

    
    face_locations = []
    face_encodings = []
    face_names = []
    process_this_frame = True

    while True:
        # Grab a single frame of video
        ret, frame = video_capture.read()
        if process_this_frame:
            small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
            rgb_small_frame = small_frame[:, :, ::-1]
            
            face_locations = face_recognition.face_locations(rgb_small_frame)
            face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)
            face_names = []
            for face_encoding in face_encodings:
                matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
                name = "Unknown"

                face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
                best_match_index = np.argmin(face_distances)
                if matches[best_match_index]:
                    name = known_face_names[best_match_index]
                face_names.append(name)
        process_this_frame = not process_this_frame

        for (top, right, bottom, left), name in zip(face_locations, face_names):
            # Scale back up face locations since the frame we detected in was scaled to 1/4 size
            top *= 4
            right *= 4
            bottom *= 4
            left *= 4
            cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)
            cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 0, 255), cv2.FILLED)
            font = cv2.FONT_HERSHEY_DUPLEX
            cv2.putText(frame, name, (left + 6, bottom - 6), font, 1.0, (255, 255, 255), 1)
            if (name!='Unknown'):
                #print(name)
                if name not in present_names:
                    print(name)
                    c=names.index(name)
                    print(c)
                    #print(regno[c])
                    present_regno.append(regno[c])
                    present_names.append(name)
                    lastpresdata[0]=str(name);
                    lastpresdata[1]=str(regno[c])

                    #p5= multiprocessing.Process(target=excel_update,args=(int(regno[c])))
                    #p5.start()
                    excel_update(int(regno[c]))

        cv2.imshow('Video', frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    # Release handle to the webcam
    video_capture.release()
    cv2.destroyAllWindows()


def DUAL_MODE(present_names,present_regno):
    rfidname=''
    print("DUAL MODE IS NOW INITIALIZING")
    print("INITIALIZING RFID LISTENER")
    print("...............")
    arduino = serial.Serial(port,115200, timeout=.1)
    print("RFID INITIALIZED")
    print("INITIALIZING CAMERA\n.\n..\n...\n....\n.....\n")
    video_capture = cv2.VideoCapture(cam_id)
    print("CAMERA INITIALIZED SUCCESSFULLY !!")

    face_locations = []
    face_encodings = []
    face_names = []
    process_this_frame = True

    while True:
        # Grab a single frame of video
        ret, frame = video_capture.read()
        if process_this_frame:
            small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
            rgb_small_frame = small_frame[:, :, ::-1]
            
            face_locations = face_recognition.face_locations(rgb_small_frame)
            face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)
            face_names = []
            for face_encoding in face_encodings:
                matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
                name = "Unknown"

                face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
                best_match_index = np.argmin(face_distances)
                if matches[best_match_index]:
                    name = known_face_names[best_match_index]
                face_names.append(name)
        process_this_frame = not process_this_frame
        d=None
        data = arduino.readline().decode('utf8')
        if data:
            for i in range(0,len(RFID)):
                if int(data)==int(RFID[i]):
                    print(names[i])
                    d=names[i]
                    break
        for (top, right, bottom, left), name in zip(face_locations, face_names):
            # Scale back up face locations since the frame we detected in was scaled to 1/4 size
            top *= 4
            right *= 4
            bottom *= 4
            left *= 4
            cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)
            cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 0, 255), cv2.FILLED)
            font = cv2.FONT_HERSHEY_DUPLEX
            cv2.putText(frame, name, (left + 6, bottom - 6), font, 1.0, (255, 255, 255), 1)
            if (name!='Unknown'):
                if name not in present_names:
                    c=names.index(name)
                    if names[c] == d:
                        print("fgfg")
                    #print(regnum[c])
                        present_regno.append(regno[c])
                        present_names.append(name)
                        lastpresdata[0]=str(name)
                        lastpresdata[1]=str(regno[c])
                        excel_update(int(regno[c]))

        cv2.imshow('Video', frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    # Release handle to the webcam
    video_capture.release()
    cv2.destroyAllWindows()


def RFID_MODE(present_names,present_regno):

    print("STARTING RFID MODE")
    print(".\n..\n...\n....\n")
    print("RUNNING")
    print(RFID)
    print(names)
    #arduino = serial.Serial(port,115200, timeout=.1)
    while True:
        data = arduino.readline()
        data=data.decode('utf-8')
        print(data)
        if data:
            if data in RFID:
                i = RFID.index(data)
                if names[i] not in present_names:
                    present_names.append(names[i])
                    print(names[i])
                    present_regno.append(regno[i])
                    lastpresdata[0]=str(names[i])
                    lastpresdata[1]=str(regno[i])
                    excel_update(int(regno[i]))
                    #finish_all()

def refreshgui():

    global ab ,no , percent

    absupdate()
    presupdate()
    lastpresupdater()

    ab=len(names)-len(present_names)
    no=len(present_names)

    tt=str(ab)+str(no)

    percent=round((no/len(names))*100, 2)

    progressbar.set(percent/100)

    #arduino.write((str(tt)+"\r\n").encode("utf-8"))
    #print(arduino.readline().decode("utf-8"))


    label.configure(text="NO.OF STUDENTS PRESENT="+str(no))
    label_1.configure(text="NO.OF STUDENTS ABSENT="+str(ab))
    label_ratio.configure(text=str(no)+"/"+str(len(names)))
    label_percent.configure(text=str(percent)+"%")

    frame_1.after(5000,refreshgui)



app = customtkinter.CTk()
app.geometry("1920x1080")
app.title("ATTENDANCE SYSTEM GUI")

started=False

def modes(mode):
    optionmenu_1.set(mode)
    segemented_button.set(mode)

    global lastmode
    global started

    if mode == "VISION MODE":
        print("llol")
        if started == True:
            lastmode.kill()

        p2=multiprocessing.Process(target=visionmode,args=(present_names,present_regno))
        started=True
        lastmode=p2
        p2.start()
    elif mode == "RFID MODE":
        if started == True:
            lastmode.kill()
        print("rfid")
        p3=multiprocessing.Process(target=RFID_MODE,args=(present_names,present_regno))
        lastmode=p3
        started=True;
        p3.start()
        #RFID_MODE()
    elif mode == "DUAL MODE":
        if started == True:
            lastmode.kill()
        p4=multiprocessing.Process(target=DUAL_MODE,args=(present_names,present_regno))
        lastmode=p4
        started=True
        p4.start()
        print("dual")
        #DUAL_MODE()


def button_callback():
    global absent_names

    print("button")
    new_window = customtkinter.CTkToplevel(app)
    new_window.title("ABSENT LIST")
    new_window.geometry("350x500")
    textbox3= customtkinter.CTkTextbox(new_window,fg_color="transparent" , font=("default",20),wrap="none" , width=350,height=500)
    textbox3.grid(row=0, column=1)
    textbox3.insert("0.1","ABSENT students reg.no\n")
    absent_names= list((Counter(names)-Counter(present_names)).elements())
    absent_regno= list((Counter(regno)-Counter(present_regno)).elements())

    for i in range(len(regno)):
        textbox3.insert("2.0",str(absent_regno [i])+ "  ---->  " +absent_names[i]+"\n" )

    textbox3.configure(state="disabled")

rfdata=None

def add_new_entry():
    def exit():
        nam=entry.get().upper()
        regn=regentry.get().upper()

        datas=openpyxl.load_workbook("/home/mythilicharan/Downloads/datas.xlsx")
        data1 = datas.active
        maxc=data1.max_row+1
        
        data1.cell(maxc,1).value=nam
        
        data1.cell(maxc,3).value=str(rfdata)
        data1.cell(maxc,2).value=regn

        datas.save(filename=outpath+"datas.xlsx")

        print(nam,regn,rfdata)
        add_new.destroy()

    def startcapture():
        cam = cv2.VideoCapture(cam_id)
        while True:
            ret, frame = cam.read()
            cv2.imshow('frame',frame)
            k = cv2.waitKey(1)
            if cv2.waitKey(1) & 0xFF == ord('q'):
                break
            if k%256 == 32:
                # SPACE pressed
                cv2.imwrite(str(entry.get().upper())+".png", frame)
                print("captures")

                break
        cam.release()
        capt = customtkinter.CTkLabel(master=add_new, text="SUCCESSFULLY CAPTURED\n ", font=("default",20))
        capt.place(relx=0.60, rely=0.18, anchor="nw")
        cv2.destroyAllWindows()
        print("captured")

    def scannow():

        global rfdata

        arduino = serial.Serial(port,115200, timeout=.1)
        while True:
            data = arduino.readline().decode('utf8')
            if data:
                #if data not in RFID:
                rfdata=data
                print(rfdata)
                break;

        scanrfidbutton = customtkinter.CTkButton(master=add_new,text="SCANNED !")
        scanrfidbutton.pack(pady=10, padx=10)
        scanrfidbutton.place(relx=0.60, rely=0.26, anchor="nw")
        print (rfdata)



    add_new = customtkinter.CTkToplevel(app)
    add_new.title("Add new entry")
    add_new.geometry("700x500")
    newlabel = customtkinter.CTkLabel(master=add_new, text="ENTER YOUR NAME : ", font=("default",20))
    newlabel.place(relx=0.33, rely=0.05, anchor="e")

    entry = customtkinter.CTkEntry(master=add_new,
                               placeholder_text="name (CAPS ONLY)",
                               width=200,
                               height=30,
                               border_width=2,corner_radius=5)
    entry.place(relx=0.38, rely=0.025, anchor="nw")


    reglabel = customtkinter.CTkLabel(master=add_new, text="ENTER REG,NO : ", font=("default",20))
    reglabel.place(relx=0.33, rely=0.13, anchor="e")

    regentry = customtkinter.CTkEntry(master=add_new,
                               placeholder_text="reg.no",
                               width=200,
                               height=30,
                               border_width=2,corner_radius=5)

    regentry.place(relx=0.38, rely=0.1, anchor="nw")

    spst = customtkinter.CTkLabel(master=add_new, text="TAKE SNAPSHOT : ", font=("default",20))
    spst.place(relx=0.33, rely=0.21, anchor="e")

    spstbutton = customtkinter.CTkButton(master=add_new, command=startcapture, text="START CAMERA")
    spstbutton.pack(pady=10, padx=10)
    spstbutton.place(relx=0.38, rely=0.18, anchor="nw")

    capt = customtkinter.CTkLabel(master=add_new, text="press space bar on \n the window to capture", font=("default",20))
    capt.place(relx=0.60, rely=0.18, anchor="nw")

    getid = customtkinter.CTkLabel(master=add_new, text="SCAN RFID : ", font=("default",20))
    getid.place(relx=0.33, rely=0.29, anchor="e")

    rfidbutton = customtkinter.CTkButton(master=add_new, command=scannow, text="SCANNOW")
    rfidbutton.pack(pady=10, padx=10)
    rfidbutton.place(relx=0.38, rely=0.26, anchor="nw")

    finishbutton = customtkinter.CTkButton(master=add_new, command=exit ,text="FINISH",width=300)
    finishbutton.pack(pady=10, padx=10)
    finishbutton.place(relx=0.38, rely=0.4, anchor="nw")


frame_1 = customtkinter.CTkFrame(master=app)
frame_1.pack(pady=20, padx=60, fill="both", expand=True)


optionmenu_1 = customtkinter.CTkOptionMenu(frame_1,command=modes,width=300,height=50 ,font =("default",30) ,values=["VISION MODE", "RFID MODE", "DUAL MODE"])
optionmenu_1.pack(pady=10, padx=10)
optionmenu_1.set("SELECTMODE")
optionmenu_1.place(relx=0.128, rely=0.05, anchor="center")


button_1 = customtkinter.CTkButton(master=frame_1, command=button_callback , text="SHOW ABSENTEES REG.NO")
button_1.pack(pady=10, padx=10)
button_1.place(relx=0.9, rely=0.5, anchor="center")

button_2 = customtkinter.CTkButton(master=frame_1, command=add_new_entry , text="          ADD NEW ENTRY          ")
button_2.pack(pady=10, padx=10)
button_2.place(relx=0.9, rely=0.55, anchor="center")

no=15
ab=15
label = customtkinter.CTkLabel(master=frame_1, text="NO.OF STUDENTS PRESENT="+str(no) , font=("default",20))
label.place(relx=0.98, rely=0.05, anchor="e")

label_1 = customtkinter.CTkLabel(master=frame_1, text="NO.OF STUDENTS ABSENT="+str(ab) , font=("default",20))
label_1.place(relx=0.98, rely=0.1, anchor="e")

label_ratio = customtkinter.CTkLabel(master=frame_1, text=str(no)+"/"+str(len(names)), font=("default",80))
label_ratio.place(relx=0.98, rely=0.3, anchor="e")

label_percent = customtkinter.CTkLabel(master=frame_1, text=str(no)+"%", font=("default",55))
label_percent.place(relx=0.98, rely=0.41, anchor="e")

label_date = customtkinter.CTkLabel(master=frame_1, text=str(today_date) , font=("default",20))
label_date.place(relx=0.98, rely=0.15, anchor="e")

label_month = customtkinter.CTkLabel(master=frame_1, text=str(today_name)+"  -  "+ str(this_month) , font=("default",20))
label_month.place(relx=0.98, rely=0.2, anchor="e")


segemented_button = customtkinter.CTkSegmentedButton(master=frame_1,values=["VISION MODE", "RFID MODE", "DUAL MODE"],command=modes , font = ("default",15))
segemented_button.pack(padx=20, pady=10)
segemented_button.place(relx=0.015 , rely =0.95 , anchor="w")
def change_appearance_mode_event(mode):
    customtkinter.set_appearance_mode(mode)
    
label_theme = customtkinter.CTkLabel(master=frame_1, text="SELECT THEME", font=("default",15))
label_theme.place(relx=0.13, rely=0.85, anchor="e")

appearance_mode_optionemenu = customtkinter.CTkOptionMenu(frame_1, values=["DARK" , "LIGHT"],font=("default",15),command=change_appearance_mode_event)
appearance_mode_optionemenu.place(relx=0.13, rely=0.9, anchor="e")

textbox = customtkinter.CTkTextbox(frame_1 , font=("default",18))
textbox.grid(row=0, column=0)
textbox.insert("0.1",("present students\n"))
textbox.place(relx=0.8 , rely=0.82 , anchor="e")
#textbox.configure(state="disabled")


textbox_abs = customtkinter.CTkTextbox(frame_1, font=("default",18))
textbox_abs.grid(row=0, column=0)
textbox_abs.insert("0.1","absent students")
textbox_abs.place(relx=0.98 , rely=0.82 , anchor="e")
#textbox_abs.configure(state="disabled")

progressbar = customtkinter.CTkProgressBar(master=frame_1,width=300,height=15)
progressbar.pack(padx=20, pady=10)
progressbar.place(relx=0.128, rely=0.2, anchor="center")

intro_label = customtkinter.CTkLabel(master=frame_1, text="USER INTERFACE\n FOR ATTENDANCE SYSTEM \n ( DEPT.OF MECHATRONICS ENGINEERING )", font=("default",25))
intro_label.place(relx=0.5, rely=0.1, anchor="center")

def initialize():

    print("INITIALIZING PLEASE WAIT....")
    daily=openpyxl.load_workbook(day_excel_path,data_only=True)
    monthly=openpyxl.load_workbook(month_excel_path,data_only=True)
    global today_column
    global max_row
    day=daily.active
    month=monthly.active
    today_column=month['A26'].value

    for k in range(min_row,80):
        val=day.cell(row=k,column=2)
        val1=val.value
        print(val1)
        if(val1 is None):
            max_row=k
            print(max_row)
            break
    
    for i in range(min_row,max_row):
        name_cell=day.cell(row=i,column=2)
        reg_cell=day.cell(row=i,column=1)
        uid_cell=day.cell(row=i,column=6)

        names.append(name_cell.value)
        regno.append(reg_cell.value)
        RFID.append(uid_cell.value)

        #path_names.append(img_path+name_cell.value+".jpg")


    #for i in range(0,len(names)):
    #    known_face_encodings.append(face_recognition.face_encodings(face_recognition.load_image_file(path_names[i]))[0])
    #    known_face_names.append(names[i])
    daily1=openpyxl.load_workbook(day_excel_path)
    monthly1=openpyxl.load_workbook(month_excel_path)

    day1=daily.active
    month1=monthly.active

    daily1.save(filename=outpath+today_date+today_name+"11.xlsx")
    monthly1.save(filename=outpath+this_month+"111.xlsx")

    print("INITIALIZED...")

    refreshgui()

    button_i = customtkinter.CTkButton(master=frame_1,  text="INITIALIZED")
    button_i.pack(pady=10, padx=10)
    button_i.place(relx=0.5, rely=0.95, anchor="center")
    print(names,regno,RFID,max_row,today_column)

def button_event():
    print("hello")

def finish_all():
    print("finished")
    prmopt = customtkinter.CTkButton(master=frame_1,text="FINISHED !",font=("default",20))
    prmopt.pack(pady=10, padx=10)
    prmopt.place(relx=0.5, rely=0.65, anchor="center")

def mailnow():
    print(checkbox_1.get())
    yag.send('mythilicharanr@gmail.com', 'daily record', contents)
    print("mailed successfully")

def presupdate():
    textbox.configure(state="normal")
    textbox.delete("0.0","end")
    textbox.insert("0.1","present students")
    for i in range(len(present_names)):
        textbox.insert("1.0",present_names[i]+"\n")
    textbox_abs.configure(state="disabled")


def absupdate():
    textbox_abs.configure(state="normal")
    textbox_abs.delete("0.0","end")
    global absent_names
    textbox_abs.insert("0.1","absent students")
    absent_names= list((Counter(names)-Counter(present_names)).elements())
    for i in range(len(absent_names)):
        textbox_abs.insert("1.0",absent_names[i]+"\n")
    textbox_abs.configure(state="disabled")


def lastpresupdater():
    if len(lastpresdata)==4:
        new_text=lastpresdata[0]+"\n"+lastpresdata[1]+"\n"+lastpresdata[2]+"\n"+lastpresdata[3]
        datalabel.configure(text=new_text,font=("default",20),fg_color="transparent")
        #datalabel.place(relx=0.18, rely=0.7, anchor="e")







finish_allbutton = customtkinter.CTkButton(master=frame_1, command=finish_all , text="FINISH ATTENDANCE",font=("default",20))
finish_allbutton.pack(pady=10, padx=10)
finish_allbutton.place(relx=0.5, rely=0.55, anchor="center")

mailbutton = customtkinter.CTkButton(master=frame_1, command=mailnow , text="UPDATE DATABASE AND SEND MAIL",font=("default",20))
mailbutton.pack(pady=10, padx=10)
mailbutton.place(relx=0.5, rely=0.6, anchor="center")

lastpres= customtkinter.CTkLabel(master=frame_1, text="Last entry data", font=("default",20))
lastpres.place(relx=0.1, rely=0.28, anchor="center")

lastdet= customtkinter.CTkLabel(master=frame_1, text=" NAME :\n REG.NO :\n PRESENT TIME :\n MONTH'S att. % :\n (till today)", font=("default",20), justify ="left")
lastdet.place(relx=0.18, rely=0.7, anchor="e")

datalabel=customtkinter.CTkLabel(master=frame_1, text="none\nnone\nnone\nnone\n", font=("default",20), justify ="right")
datalabel.place(relx=0.28, rely=0.68, anchor="e")

my_image = customtkinter.CTkImage(light_image=Image.open("/home/mythilicharan/Downloads/png-clipart-user-icon-foreigners-avatar-child-face.jpeg"),
                                  size=(200,150))
buttonz = customtkinter.CTkButton(frame_1,text='', image=my_image ,fg_color="transparent" , hover="false")
buttonz.place(relx=0.1 , rely=0.45 , anchor="center")


#button_i = customtkinter.CTkButton(master=frame_1, command=initialize , text="INITIALIZE")
#button_i.pack(pady=10, padx=10)
#button_i.place(relx=0.5, rely=0.9, anchor="center")


def cameraid(idc):
    global cam_id
    print(idc)
    cam_id=idc

def serialport(portname):
    global port
    port=portname
    print(portname)


week_label1 = customtkinter.CTkLabel(master=frame_1, text="SELECT CAMERA INDEX AND SERIAL PORT\n (already set to defaults)", font=("default",20))
week_label1.place(relx=0.5, rely=0.35, anchor="center")

optionmenu_6 = customtkinter.CTkOptionMenu(frame_1,command=cameraid,font =("default",20) ,values=["0","1","2"])
optionmenu_6.pack(pady=10, padx=10)
optionmenu_6.set("0")
optionmenu_6.place(relx=0.4, rely=0.45, anchor="center")

optionmenu_5 = customtkinter.CTkOptionMenu(frame_1,command=serialport,font =("default",20) ,values=['/dev/ttyUSB0','/dev/ttyUSB1','/dev/ttyUSB2'])
optionmenu_5.pack(pady=10, padx=10)
optionmenu_5.set('/dev/ttyUSB0')
optionmenu_5.place(relx=0.6, rely=0.45, anchor="center")




print(cam_id,port)



checkbox_1 = customtkinter.CTkCheckBox(master=frame_1,text="INCLUDE MONTHLY REPORT")
checkbox_1.pack(pady=10, padx=10)
checkbox_1.place(relx=0.5, rely=0.65, anchor="center")


checkbox_2 = customtkinter.CTkCheckBox(master=frame_1,text="MARK ABSENTEES AS  'A'      ")
checkbox_2.pack(pady=10, padx=10)
checkbox_2.place(relx=0.5, rely=0.70, anchor="center")

refreshbutton = customtkinter.CTkButton(master=frame_1, command=refreshgui , text="REFRESH!",font=("default",20))
refreshbutton.pack(pady=10, padx=10)
refreshbutton.place(relx=0.5, rely=0.8, anchor="center")

initialize()
#refreshgui()


app.mainloop()
